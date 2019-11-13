import asyncio
import aiohttp
from bs4 import BeautifulSoup
import ssl
import time
import openpyxl
from fake_useragent import UserAgent
import random
import concurrent
from pypeln import TaskPool
import logging

from free_proxy_scraper import get_us_proxy_list

logging.basicConfig(filename='app.log', 
                    level=logging.INFO,
                    filemode='a', ## append ##
                    format='[%(asctime)s] - %(name)s - %(levelname)s - %(message)s',
                    datefmt='%H:%M:%S')

def timeit(func):
    '''
    Decorator to time it takes for function to return. Outputs to stdout.
    '''
    def wrapper(*args, **kwargs):
        start_time = time.time()
        res = func(*args, **kwargs)
        print("--- %s seconds ---" % (time.time() - start_time))
        return res
    return wrapper

class TaskPool(object):
    '''
    Reference: https://medium.com/@cgarciae/making-an-infinite-number-of-requests-with-python-aiohttp-pypeln-3a552b97dc95
    '''
    def __init__(self, workers):
        self._semaphore = asyncio.Semaphore(workers)
        self._tasks = set()
        

    async def put(self, coro):
        await self._semaphore.acquire()

        task = asyncio.ensure_future(coro)
        self._tasks.add(task)
        task.add_done_callback(self._on_task_done)

    def _on_task_done(self, task):
        self._tasks.remove(task)
        self._semaphore.release()

    async def join(self):
        await asyncio.gather(*self._tasks)

    async def __aenter__(self):
        return self

    def __aexit__(self, exc_type, exc, tb):
        return self.join()

async def parse_html(url, i, hdr, user_agent, ip_list, session):
    '''
    Reference (user_agent) : https://pypi.org/project/user-agents/
    Reference (aiohttp) : https://aiohttp.readthedocs.io/en/stable/client_reference.html

    Function
    ---------------------
    Sets up async http requests.
    ---------------------

    Inputs
    ---------------------
    urls : list of URLs
    ssl_ctx : ssl connection param
    hdr : used to “spoof” the User-Agent header value, which is used by a browser to identify itself
    user_agent : python object used to generate random user agent
    ip_list : list of free HTTP proxy IP addresses in the US
    ---------------------
    '''
    hdr['User-Agent'] = user_agent.random
    # ip = random.choice(ip_list) ## free ip source unreliable ##

    try:
        # async with session.get(url, headers=hdr, proxy=ip) as response: ## free ip source unreliable ##
        async with session.get(url, headers=hdr) as response:
            html = await response.text()
    except aiohttp.client_exceptions.ServerDisconnectedError as e:
        # if ip in ip_list: ip_list.remove(ip) ## remove bad proxy ips to prevent reuse ## ## free ip source unreliable ##
        logging.exception(f'[{i}] - {e}')
        
        return await parse_html(url, i, hdr, user_agent, ip_list, session)

    if html:
        soup_main = BeautifulSoup(html, 'html.parser')

        txt = soup_main.find('td').text.strip().replace(',', '')

        write_to_excel(txt, i)

def write_to_excel(txt, i):
    '''
    Function
    ---------------------
    Writes counter and first parse 'td' text to excel.
    ---------------------

    Inputs
    ---------------------
    txt : first parsed 'td' tag text
    i : counter
    ---------------------
    '''
    fn = 'temp.xlsx'
    wb = openpyxl.load_workbook(filename=fn)
    ws = wb['Sheet1']
    row = ws.max_row + 1

    ws.cell(row=row, column=1, value=i)
    ws.cell(row=row, column=2, value=txt)

    wb.save(fn)
    wb.close()

    print(f'Wrote {i}')

async def async_setup(urls, ssl_ctx, hdr, user_agent, ip_list=[]):
    '''
    Reference (asyncio.Semaphore) : https://docs.python.org/3/library/asyncio-sync.html#asyncio.Semaphore
    Reference (aiohttp) : https://aiohttp.readthedocs.io/en/stable/client_reference.html

    Function
    ---------------------
    Sets up async http requests.
    ---------------------

    Inputs
    ---------------------
    urls : list of URLs
    ssl_ctx : ssl connection param
    hdr : used to “spoof” the User-Agent header value, which is used by a browser to identify itself
    user_agent : python object used to generate random user agent
    ip_list : list of free HTTP proxy IP addresses in the US
    ---------------------
    '''
    limit = 10 ## limit the number of threads that can be run at any one time ##

    async with asyncio.Semaphore(limit):
        async with aiohttp.TCPConnector(ssl=ssl_ctx, limit=None) as connector:
            async with aiohttp.ClientSession(connector=connector) as session, TaskPool(limit) as tasks:
                for i in range(len(urls)):
                    await tasks.put(parse_html(urls[i], i, hdr, user_agent, ip_list, session))

if __name__ == '__main__':
    urls = []
    for i in range(1, 20):
        for j in range(1, 10):
            urls.append(f'http://testing-ground.scraping.pro/table?products=1&years={i}&quarters={j}') ## free web scraping site ##

    ssl_ctx = ssl.create_default_context()
    ssl_ctx.check_hostname = False
    ssl_ctx.verify_mode = ssl.CERT_NONE

    hdr = {'User-Agent': '',
           'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
           'Accept-Charset': 'ISO-8859-1,utf-8;q=0.7,*;q=0.3',
           'Accept-Encoding': 'none',
           'Accept-Language': 'en-US,en;q=0.8',
           'Connection': 'keep-alive'}

    user_agent = UserAgent(fallback='Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.11 (KHTML, like Gecko) Chrome/23.0.1271.64 Safari/537.11')
    # ip_list = get_us_proxy_list(http_only=True) ## free ip source unreliable ##

    # --------------------------------------------- #
    # run the event loop until all tasks are complete
    loop = asyncio.get_event_loop()
    start_time = time.time()
    loop.run_until_complete(async_setup(urls, ssl_ctx, hdr, user_agent))
    print("--- %s seconds ---" % (time.time() - start_time))
    print('Done.')
    # --------------------------------------------- #